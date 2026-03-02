import streamlit as st
import pytesseract
from PIL import Image
import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io
import re

# Dados das progressões inline
PROGRESSOES_PISTA_TRILHA = {
    1: {"area": "Físico", "desc": "Participar de pelo menos cinco atividades de patrulha ao ar livre"},
    2: {"area": "Físico", "desc": "Conhecer e aplicar normas de limpeza no tratamento de alimentos"},
    3: {"area": "Físico", "desc": "Aferir seu passo duplo e conhecer medidas do corpo"},
    4: {"area": "Físico", "desc": "Conhecer elementos da Caixa de Primeiros Socorros"},
    5: {"area": "Físico", "desc": "Conhecer ações iniciais em acidentes e cuidar de ferimentos"},
    6: {"area": "Físico", "desc": "Aplicar medidas de segurança nas atividades"},
    7: {"area": "Físico", "desc": "Prevenir males da exposição ao sol"},
    8: {"area": "Físico", "desc": "Manter hábitos de higiene e cuidado com uniforme"},
    9: {"area": "Físico", "desc": "Classificar lixo e tratar resíduos"},
    10: {"area": "Físico", "desc": "Participar da manutenção do canto de patrulha"},
    11: {"area": "Físico", "desc": "Montar mochila para acampamento de 3 dias"},
    12: {"area": "Físico", "desc": "Montar cardápio com refeições equilibradas"},
    13: {"area": "Físico", "desc": "Colaborar na elaboração de alimentos"},
    14: {"area": "Físico", "desc": "Montar solução para purificação de água"},
    15: {"area": "Físico", "desc": "Utilizar diversos tipos de fogos"},
    16: {"area": "Físico", "desc": "Organizar tempo com agenda"},
    17: {"area": "Físico", "desc": "Realizar tarefas escolares no prazo"},
    18: {"area": "Físico", "desc": "Frequentar atividades da patrulha"},
    19: {"area": "Físico", "desc": "Realizar atividade física regularmente"},
    20: {"area": "Físico", "desc": "Participar de jogos respeitando regras"},
    21: {"area": "Intelectual", "desc": "Traçar e seguir sinais de pista"},
    22: {"area": "Intelectual", "desc": "Utilizar mapa e bússola"},
    23: {"area": "Intelectual", "desc": "Aplicar técnicas de tocaia"},
    24: {"area": "Intelectual", "desc": "Explorar comunidade identificando problemas"},
    25: {"area": "Intelectual", "desc": "Estimar altura e distâncias"},
    26: {"area": "Intelectual", "desc": "Ler livro e apresentar resumo"},
    27: {"area": "Intelectual", "desc": "Utilizar técnica de previsão do tempo"},
    28: {"area": "Intelectual", "desc": "Participar de Jogos Democráticos"},
    29: {"area": "Intelectual", "desc": "Participar do Conselho de Patrulha"},
    30: {"area": "Intelectual", "desc": "Participar do planejamento de excursão"},
    31: {"area": "Intelectual", "desc": "Avaliar atividades com patrulha"},
    32: {"area": "Intelectual", "desc": "Utilizar especialidades conquistadas"},
    33: {"area": "Intelectual", "desc": "Ajudar escoteiro a conquistar especialidade"},
    34: {"area": "Intelectual", "desc": "Participar de fogo de conselho"},
    35: {"area": "Intelectual", "desc": "Construir instrumento musical"},
    36: {"area": "Intelectual", "desc": "Conhecer e cantar canções tradicionais"},
    37: {"area": "Intelectual", "desc": "Ler e escrever em código secreto"},
    38: {"area": "Intelectual", "desc": "Utilizar rádio comunicador"},
    39: {"area": "Intelectual", "desc": "Montar blog ou projeto de comunicação"},
    40: {"area": "Intelectual", "desc": "Participar da construção de pioneirias"},
    41: {"area": "Intelectual", "desc": "Utilizar ferramentas de patrulha"},
    42: {"area": "Intelectual", "desc": "Participar da construção de Fogão Suspenso"},
    43: {"area": "Caráter", "desc": "Propor objetivos para melhorar vida"},
    44: {"area": "Caráter", "desc": "Participar da avaliação de progressão"},
    45: {"area": "Caráter", "desc": "Avaliar desempenho nos cargos"},
    46: {"area": "Caráter", "desc": "Explicar Lei e Promessa Escoteira"},
    47: {"area": "Caráter", "desc": "Participar de cerimônias cívicas"},
    48: {"area": "Caráter", "desc": "Explicar o que é ser leal"},
    49: {"area": "Caráter", "desc": "Aplicar conceito de lealdade"},
    50: {"area": "Caráter", "desc": "Participar como animador"},
    51: {"area": "Caráter", "desc": "Conhecer histórias de superação"},
    52: {"area": "Caráter", "desc": "Respeitar decisões do Conselho"},
    53: {"area": "Caráter", "desc": "Melhorar organização do Conselho"},
    54: {"area": "Caráter", "desc": "Participar da eleição do Monitor"},
    55: {"area": "Afetivo", "desc": "Pesquisar malefícios de drogas"},
    56: {"area": "Afetivo", "desc": "Contribuir com Livro de Patrulha"},
    57: {"area": "Afetivo", "desc": "Participar de turno de ronda"},
    58: {"area": "Afetivo", "desc": "Registrar momentos pessoais"},
    59: {"area": "Afetivo", "desc": "Participar de debate sobre filme"},
    60: {"area": "Afetivo", "desc": "Participar de Assembleias"},
    61: {"area": "Afetivo", "desc": "Propor temas para debate"},
    62: {"area": "Afetivo", "desc": "Avaliar acampamento"},
    63: {"area": "Afetivo", "desc": "Auxiliar novo integrante"},
    64: {"area": "Afetivo", "desc": "Convidar patrulha para reunião"},
    65: {"area": "Afetivo", "desc": "Participar de atividades de igualdade"},
    66: {"area": "Afetivo", "desc": "Compartilhar tarefas domésticas"},
    67: {"area": "Afetivo", "desc": "Investigar sobre mulheres na história"},
    68: {"area": "Afetivo", "desc": "Participar de cerimônia com pais"},
    69: {"area": "Afetivo", "desc": "Participar de atividade com família"},
    70: {"area": "Afetivo", "desc": "Solicitar ajuda de familiares"},
    71: {"area": "Social", "desc": "Investigar defensores de direitos humanos"},
    72: {"area": "Social", "desc": "Participar de atividades sobre Direitos Humanos"},
    74: {"area": "Social", "desc": "Colaborar para definir metas"},
    75: {"area": "Social", "desc": "Assumir cargo na patrulha"},
    76: {"area": "Social", "desc": "Participar das decisões do Conselho"},
    77: {"area": "Social", "desc": "Participar de Assembleia de Tropa"},
    78: {"area": "Social", "desc": "Estudar organização do Escotismo"},
    79: {"area": "Social", "desc": "Conhecer estrutura do Grupo"},
    80: {"area": "Social", "desc": "Realizar boas ações"},
    81: {"area": "Social", "desc": "Participar de MUTCOM"},
    82: {"area": "Social", "desc": "Fazer croqui da área de residência"},
    83: {"area": "Social", "desc": "Conhecer serviços públicos"},
    84: {"area": "Social", "desc": "Participar de comemoração regional"},
    85: {"area": "Social", "desc": "Participar de Jantar Festivo"},
    86: {"area": "Social", "desc": "Pesquisar jogos típicos"},
    87: {"area": "Social", "desc": "Participar de evento cívico"},
    88: {"area": "Social", "desc": "Explicar significado da Flor de Lis"},
    89: {"area": "Social", "desc": "Conhecer história do Grupo"},
    90: {"area": "Social", "desc": "Participar de atividade distrital"},
    91: {"area": "Social", "desc": "Participar de JOTI ou JOTA"},
    92: {"area": "Social", "desc": "Participar de atividade pela paz"},
    93: {"area": "Social", "desc": "Pesquisar sobre defensores da paz"},
    94: {"area": "Social", "desc": "Participar de projeto ambiental"},
    95: {"area": "Social", "desc": "Realizar levantamento de pegadas"},
    96: {"area": "Social", "desc": "Participar de excursão ecológica"},
    97: {"area": "Espiritual", "desc": "Fazer orações rotineiras"},
    98: {"area": "Espiritual", "desc": "Participar de celebrações religiosas"},
    99: {"area": "Espiritual", "desc": "Realizar reflexões em acampamentos"},
    100: {"area": "Espiritual", "desc": "Participar de serviço comunitário religioso"},
    101: {"area": "Espiritual", "desc": "Aplicar ensinamentos religiosos"},
    102: {"area": "Espiritual", "desc": "Apresentar relato de ensinamentos"},
    103: {"area": "Espiritual", "desc": "Participar da construção de espaço de reflexão"},
    104: {"area": "Espiritual", "desc": "Orar com oração da Tropa"},
    105: {"area": "Espiritual", "desc": "Praticar a oração"},
    106: {"area": "Espiritual", "desc": "Organizar livreto de orações"},
    107: {"area": "Espiritual", "desc": "Conhecer confissões religiosas dos amigos"},
    108: {"area": "Espiritual", "desc": "Pesquisar sobre confissão religiosa diferente"}
}

PROGRESSOES_RUMO_TRAVESSIA = {
    1: {"area": "Físico", "desc": "Participar de pelo menos 5 atividades ao ar livre da patrulha"},
    2: {"area": "Físico", "desc": "Saber explicar mudanças no corpo e males de Anorexia/Bulimia"},
    3: {"area": "Físico", "desc": "Participar de Jornada de Travessia"},
    4: {"area": "Físico", "desc": "Reconhecer animais venenosos e peçonhentos"},
    5: {"area": "Físico", "desc": "Manter Caixa de Primeiros Socorros em dia"},
    6: {"area": "Físico", "desc": "Aplicar medidas gerais de segurança em acidentes"},
    7: {"area": "Físico", "desc": "Saber agir em casos de hemorragia"},
    8: {"area": "Físico", "desc": "Participar de limpeza do Canto de Patrulha"},
    9: {"area": "Físico", "desc": "Propor e executar melhoria em local visitado"},
    10: {"area": "Físico", "desc": "Demonstrar cuidado com vestuário e costurar distintivos"},
    11: {"area": "Físico", "desc": "Montar mochila para acampamento de 5 dias"},
    12: {"area": "Físico", "desc": "Preparar 5 refeições para patrulha"},
    13: {"area": "Físico", "desc": "Montar cardápio de acampamento de fim de semana"},
    14: {"area": "Físico", "desc": "Cozinhar ao ar livre sem utensílios (comida mateira)"},
    15: {"area": "Físico", "desc": "Organizar atividades em calendário semanal"},
    16: {"area": "Físico", "desc": "Classificar atividades segundo prioridades"},
    17: {"area": "Físico", "desc": "Participar regularmente de atividades da patrulha"},
    18: {"area": "Físico", "desc": "Desenvolver passatempo ou hobbie"},
    19: {"area": "Físico", "desc": "Realizar regularmente atividade física ou esporte"},
    20: {"area": "Físico", "desc": "Participar de jogos com outros Grupos Escoteiros"},
    21: {"area": "Intelectual", "desc": "Realizar previsão do tempo por indícios naturais"},
    22: {"area": "Intelectual", "desc": "Traçar e seguir sinais de pista (1km campo ou 2km urbano)"},
    23: {"area": "Intelectual", "desc": "Orientar-se com recursos naturais, bússola e mapa"},
    24: {"area": "Intelectual", "desc": "Ler pelo menos um capítulo de Escotismo para Rapazes"},
    25: {"area": "Intelectual", "desc": "Participar de três Jogos Democráticos da Tropa"},
    26: {"area": "Intelectual", "desc": "Participar da avaliação de atividade Regional"},
    27: {"area": "Intelectual", "desc": "Explorar tema de interesse e compartilhar"},
    28: {"area": "Intelectual", "desc": "Aplicar técnicas de medição de distância ou altura"},
    29: {"area": "Intelectual", "desc": "Preparar materiais para representações artísticas"},
    30: {"area": "Intelectual", "desc": "Organizar dia de jogos na casa de companheiro"},
    31: {"area": "Intelectual", "desc": "Propor e colaborar na organização de atividades"},
    32: {"area": "Intelectual", "desc": "Organizar atividade de divulgação no colégio"},
    33: {"area": "Intelectual", "desc": "Aplicar especialidades em ações de serviço"},
    34: {"area": "Intelectual", "desc": "Ajudar jovens na conquista de especialidades"},
    35: {"area": "Intelectual", "desc": "Propor ideias de ações a serviço da comunidade"},
    36: {"area": "Intelectual", "desc": "Ser responsável por canções em Fogo de Conselho"},
    37: {"area": "Intelectual", "desc": "Organizar e participar de esquete"},
    38: {"area": "Intelectual", "desc": "Ensinar canções tradicionais do Movimento"},
    39: {"area": "Intelectual", "desc": "Construir Fogão Solar e utilizar em acampamento"},
    40: {"area": "Intelectual", "desc": "Construir chuveiro de acampamento"},
    41: {"area": "Intelectual", "desc": "Saber como funcionam serviços (telefone, internet, etc)"},
    42: {"area": "Intelectual", "desc": "Enviar e receber mensagens em morse, semáfora ou libras"},
    43: {"area": "Intelectual", "desc": "Desenhar croqui e participar de instalação de pioneirias"},
    44: {"area": "Intelectual", "desc": "Aplicar conceitos de estruturas (cavaletes, encaixes)"},
    45: {"area": "Intelectual", "desc": "Confeccionar Falcaças, Nó catau, Lais de guia"},
    46: {"area": "Intelectual", "desc": "Construir e pernoitar em abrigo natural"},
    47: {"area": "Caráter", "desc": "Propor objetivos para melhorar aspectos da vida"},
    48: {"area": "Caráter", "desc": "Participar da avaliação de progressão pessoal"},
    49: {"area": "Caráter", "desc": "Participar de reunião sobre aspectos da patrulha"},
    50: {"area": "Caráter", "desc": "Ajudar companheiro em sua progressão pessoal"},
    51: {"area": "Caráter", "desc": "Desempenhar cargo de patrulha por um ciclo"},
    52: {"area": "Caráter", "desc": "Capacitar-se para desempenhar cargo na patrulha"},
    53: {"area": "Caráter", "desc": "Avaliar desempenho nos cargos de patrulha"},
    54: {"area": "Caráter", "desc": "Participar de festival de talentos"},
    55: {"area": "Caráter", "desc": "Auxiliar companheiro a realizar Promessa Escoteira"},
    56: {"area": "Caráter", "desc": "Avaliar vivência da Promessa e Lei na Patrulha"},
    57: {"area": "Caráter", "desc": "Cantar Canção da Promessa com patrulha"},
    58: {"area": "Caráter", "desc": "Conhecer história de Caio Viana Martins"},
    59: {"area": "Caráter", "desc": "Realizar boas ações pessoais e com patrulha"},
    60: {"area": "Caráter", "desc": "Organizar Oficina de Brinquedos e doar itens"},
    61: {"area": "Caráter", "desc": "Conhecer e cantar canções para distintos momentos"},
    62: {"area": "Caráter", "desc": "Criar vídeo de campanha promovendo Grupo Escoteiro"},
    63: {"area": "Afetivo", "desc": "Demonstrar técnicas de resgate em afogamento"},
    64: {"area": "Afetivo", "desc": "Ultrapassar obstáculo com cordas (falsa baiana, etc)"},
    65: {"area": "Afetivo", "desc": "Fazer relato no Livro de Patrulha"},
    66: {"area": "Afetivo", "desc": "Conhecer e aplicar princípios de boa comunicação"},
    67: {"area": "Afetivo", "desc": "Participar de debates no Conselho e Assembleia"},
    68: {"area": "Afetivo", "desc": "Visitar outro Grupo Escoteiro"},
    69: {"area": "Afetivo", "desc": "Contribuir para manutenção do Espírito Escoteiro"},
    70: {"area": "Afetivo", "desc": "Participar de mobilização para problema social"},
    71: {"area": "Afetivo", "desc": "Saber a quem recorrer em casos de maus tratos"},
    72: {"area": "Afetivo", "desc": "Ajudar companheiro a conquistar objetivo"},
    73: {"area": "Afetivo", "desc": "Auxiliar patrulha a ter número equilibrado de meninas/meninos"},
    74: {"area": "Afetivo", "desc": "Ir ao teatro ou cinema com patrulha"},
    75: {"area": "Afetivo", "desc": "Realizar Boa Ação com membros da família"},
    76: {"area": "Afetivo", "desc": "Realizar projeto de patrulha com ajuda de pais"},
    77: {"area": "Afetivo", "desc": "Assumir responsabilidade de tarefa doméstica por 3 meses"},
    78: {"area": "Social", "desc": "Propor atividades sobre Direitos Humanos"},
    79: {"area": "Social", "desc": "Pesquisar sobre Violência Escolar e apresentar"},
    80: {"area": "Social", "desc": "Participar de atividade sobre Direitos das Crianças"},
    81: {"area": "Social", "desc": "Saber sobre Constituição e cantar Hino Nacional"},
    82: {"area": "Social", "desc": "Visitar Câmara de Vereadores"},
    83: {"area": "Social", "desc": "Saber diferenças entre poderes Legislativo/Executivo/Judiciário"},
    84: {"area": "Social", "desc": "Assistir Assembleia do Grupo Escoteiro"},
    85: {"area": "Social", "desc": "Participar de Assembleia de Tropa propondo melhorias"},
    86: {"area": "Social", "desc": "Pesquisar sobre organização do Escotismo"},
    87: {"area": "Social", "desc": "Apresentar estrutura de Grupo Escoteiro"},
    88: {"area": "Social", "desc": "Participar de mutirão/projeto Insígnia/campanha comunitária"},
    89: {"area": "Social", "desc": "Saber localização de serviços públicos"},
    90: {"area": "Social", "desc": "Participar de Safári Fotográfico na cidade"},
    91: {"area": "Social", "desc": "Identificar problemas da cidade e propor soluções"},
    92: {"area": "Social", "desc": "Pesquisar lenda brasileira para Fogo de Conselho"},
    93: {"area": "Social", "desc": "Aprender canções e danças do Brasil"},
    94: {"area": "Social", "desc": "Confeccionar artesanato típico do Brasil"},
    95: {"area": "Social", "desc": "Participar de atividade escoteira distrital/regional/nacional"},
    96: {"area": "Social", "desc": "Pesquisar sobre história do Escotismo no Brasil"},
    97: {"area": "Social", "desc": "Realizar atividade com patrulha de outro Grupo"},
    98: {"area": "Social", "desc": "Participar de JOTI ou JOTA"},
    99: {"area": "Social", "desc": "Manter contato com escoteiro de outro país/jantar/insígnia"},
    100: {"area": "Social", "desc": "Visitar organização ambiental e pesquisar problemas"},
    101: {"area": "Social", "desc": "Participar de projeto de conservação ambiental"},
    102: {"area": "Social", "desc": "Identificar pegadas ou participar da Tribo da Terra"},
    103: {"area": "Espiritual", "desc": "Participar regularmente de cultos religiosos"},
    104: {"area": "Espiritual", "desc": "Auxiliar em celebração da comunidade religiosa"},
    105: {"area": "Espiritual", "desc": "Ler pelo menos um livro sagrado"},
    106: {"area": "Espiritual", "desc": "Realizar atividades de reflexão em acampamento"},
    107: {"area": "Espiritual", "desc": "Projetar e construir lugar de oração em acampamento"},
    108: {"area": "Espiritual", "desc": "Organizar momentos de oração com patrulha/família"},
    109: {"area": "Espiritual", "desc": "Aplicar ensinamentos religiosos na vida"},
    110: {"area": "Espiritual", "desc": "Avaliar ações de acordo com ensinamentos religiosos"},
    111: {"area": "Espiritual", "desc": "Convidar patrulha para ações da comunidade religiosa"},
    112: {"area": "Espiritual", "desc": "Discutir episódio sobre fanatismo religioso"},
    113: {"area": "Espiritual", "desc": "Confeccionar calendário de festividades religiosas"}
}

EMOJIS_AREAS = {
    "Físico": "💪",
    "Intelectual": "🧠",
    "Caráter": "❤️",
    "Afetivo": "🤝",
    "Social": "👥",
    "Espiritual": "✨"
}

# Configuração
st.set_page_config(page_title="Scout Progress Validator", page_icon="🏕️", layout="wide")

st.markdown("""
<style>
    .main {padding: 2rem;}
    .stButton>button {width: 100%; background-color: #4472C4; color: white; 
                      font-weight: bold; padding: 0.75rem; border-radius: 0.5rem;}
    .stButton>button:hover {background-color: #2952A3;}
    h1 {color: #4472C4; text-align: center;}
</style>
""", unsafe_allow_html=True)

st.markdown("# 🏕️ Scout Progress Validator v4.0")
st.markdown("### Sistema de Validação - 2 Modos de OCR")
st.markdown("---")

# Session state
if 'processed' not in st.session_state:
    st.session_state.processed = False
if 'excel_data' not in st.session_state:
    st.session_state.excel_data = None
if 'stats' not in st.session_state:
    st.session_state.stats = None

# OCR Simplificado - Detecção por Círculos Verdes
def ocr_simplificado(image, level):
    """Detecta círculos verdes e números próximos"""
    try:
        max_item = 108 if level == "Pista/Trilha" else 113
        
        img_array = np.array(image)
        height, width = img_array.shape[:2]
        
        # 1. DETECTAR CÍRCULOS VERDES
        # Converter para HSV (melhor para detectar cores)
        img_hsv = cv2.cvtColor(img_array, cv2.COLOR_RGB2HSV)
        
        # Faixa de verde em HSV (mais preciso que RGB)
        # Verde: H=35-85, S=50-255, V=50-255
        lower_green = np.array([35, 50, 50])
        upper_green = np.array([85, 255, 255])
        
        mascara_verde = cv2.inRange(img_hsv, lower_green, upper_green)
        
        # Encontrar contornos dos círculos verdes
        contours, _ = cv2.findContours(mascara_verde, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        # Filtrar contornos que parecem círculos
        circulos_verdes = []
        for contour in contours:
            area = cv2.contourArea(contour)
            if 20 < area < 500:  # Tamanho razoável de círculo
                x, y, w, h = cv2.boundingRect(contour)
                aspect_ratio = float(w) / h if h > 0 else 0
                
                # Círculo tem aspect ratio próximo de 1
                if 0.7 < aspect_ratio < 1.3:
                    # Centro do círculo
                    cx = x + w // 2
                    cy = y + h // 2
                    circulos_verdes.append((cx, cy, y))
        
        # 2. EXTRAIR NÚMEROS COM OCR
        img_gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
        
        # OCR na imagem completa para pegar todos os números
        text_completo = pytesseract.image_to_string(img_gray, config='--psm 6 -l por')
        
        # Extrair números e suas posições aproximadas
        linhas = text_completo.split('\n')
        
        # Mapear número -> posição Y aproximada
        numeros_posicoes = []
        pixels_por_linha = height // max(len(linhas), 1)
        
        for i, linha in enumerate(linhas):
            # Procurar padrão: "número - descrição" ou "número. descrição"
            match = re.match(r'^\s*(\d{1,3})\s*[-\.\)]', linha)
            if match:
                num = int(match.group(1))
                if 1 <= num <= max_item:
                    y_pos = i * pixels_por_linha + pixels_por_linha // 2
                    numeros_posicoes.append((num, y_pos))
        
        # 3. ASSOCIAR CÍRCULOS VERDES COM NÚMEROS
        concluidas = []
        
        for num, num_y in numeros_posicoes:
            # Verificar se há círculo verde próximo (mesma altura ±40 pixels)
            for cx, cy, circle_y in circulos_verdes:
                if abs(num_y - circle_y) < 40:  # Tolerância de 40 pixels
                    concluidas.append(num)
                    break
        
        # Todos os números detectados
        todos = [num for num, _ in numeros_posicoes]
        
        # Pendentes = todos - concluídas
        concluidas = sorted(set(concluidas))
        pendentes = sorted(set(todos) - set(concluidas))
        
        # 4. DETECTAR DATAS (opcional)
        datas = {}
        for match in re.finditer(r'(\d+)[^\d]*(\d{2}/\d{2}/\d{4})', text_completo):
            num, data = match.groups()
            num = int(num)
            if 1 <= num <= max_item:
                datas[num] = data
        
        return {
            'concluidas': concluidas, 
            'pendentes': pendentes, 
            'datas': datas,
            'debug': {
                'circulos_encontrados': len(circulos_verdes),
                'numeros_detectados': len(todos),
                'total_esperado': max_item,
                'erro': None
            }
        }
    except Exception as e:
        # Se der erro, retorna vazio com info do erro
        return {
            'concluidas': [],
            'pendentes': [],
            'datas': {},
            'debug': {
                'circulos_encontrados': 0,
                'numeros_detectados': 0,
                'total_esperado': max_item,
                'erro': str(e)
            }
        }

# Gerar planilha simplificada
def gerar_planilha_simplificada(concluidas, pendentes, scout_name, level, datas={}):
    # Escolher banco de dados correto
    progressoes = PROGRESSOES_PISTA_TRILHA if level == "Pista/Trilha" else PROGRESSOES_RUMO_TRAVESSIA
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Progressões"
    
    verde = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    amarelo = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    bold_font = Font(bold=True)
    
    # Cabeçalho
    ws.merge_cells('A1:E1')
    ws['A1'] = f"📊 VALIDAÇÃO DE PROGRESSÃO - {scout_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = f"Nível: {level}"
    ws['A3'].font = bold_font
    ws['A4'] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A4'].font = bold_font
    
    total = len(concluidas) + len(pendentes)
    ws['A6'] = f"Total: {total} itens"
    ws['A7'] = f"✅ Concluídas: {len(concluidas)}"
    ws['A7'].fill = verde
    ws['A8'] = f"⏳ Pendentes: {len(pendentes)}"
    ws['A8'].fill = amarelo
    ws['A9'] = f"📊 % Conclusão: {(len(concluidas)/total*100 if total>0 else 0):.1f}%"
    ws['A9'].font = bold_font
    
    # Tabela
    row = 11
    headers = ['#', 'Área', 'Descrição', 'Status', 'Data']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row += 1
    
    # Concluídas
    for num in concluidas:
        info = progressoes.get(num, {"area": "?", "desc": "?"})
        ws.cell(row=row, column=1, value=num).fill = verde
        ws.cell(row=row, column=2, value=f"{EMOJIS_AREAS.get(info['area'], '')} {info['area']}").fill = verde
        ws.cell(row=row, column=3, value=info['desc']).fill = verde
        ws.cell(row=row, column=4, value="✅ Concluída").fill = verde
        ws.cell(row=row, column=5, value=datas.get(num, '')).fill = verde
        row += 1
    
    # Pendentes
    for num in pendentes:
        info = progressoes.get(num, {"area": "?", "desc": "?"})
        ws.cell(row=row, column=1, value=num).fill = amarelo
        ws.cell(row=row, column=2, value=f"{EMOJIS_AREAS.get(info['area'], '')} {info['area']}").fill = amarelo
        ws.cell(row=row, column=3, value=info['desc']).fill = amarelo
        ws.cell(row=row, column=4, value="⏳ Pendente").fill = amarelo
        row += 1
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 70
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, total, len(concluidas), len(pendentes)

# Interface
col1, col2 = st.columns([2, 1])

with col1:
    st.markdown("### 📤 Upload da Imagem")
    uploaded_file = st.file_uploader("Arraste ou clique para selecionar", type=['png', 'jpg', 'jpeg', 'webp'])
    if uploaded_file:
        image = Image.open(uploaded_file)
        st.image(image, caption="Imagem carregada", use_column_width=True)

with col2:
    st.markdown("### ⚙️ Configurações")
    
    modo_ocr = st.radio(
        "🎯 Modo de OCR",
        options=["OCR Simplificado", "OCR Completo"],
        help="Simplificado: apenas números + check (rápido e preciso) | Completo: extrai texto completo"
    )
    
    scout_name = st.text_input("Nome do Jovem", placeholder="Ex: Gustavo Santos")
    level = st.radio("Nível", options=["Pista/Trilha", "Rumo/Travessia"])
    
    st.markdown("---")
    
    if st.button("🚀 PROCESSAR", disabled=not (uploaded_file and scout_name)):
        with st.spinner("⏳ Processando..."):
            
            if modo_ocr == "OCR Simplificado":
                st.info("🎯 Modo Simplificado: Detectando números e checks...")
                resultado = ocr_simplificado(image, level)
                
                with st.expander("🔍 Detecção OCR"):
                    debug = resultado.get('debug', {})
                    if debug.get('erro'):
                        st.error(f"❌ ERRO: {debug['erro']}")
                    st.write(f"🎯 Círculos verdes encontrados: {debug.get('circulos_encontrados', '?')}")
                    st.write(f"📋 Números detectados: {debug.get('numeros_detectados', '?')}")
                    st.write(f"📊 Esperado: {debug.get('total_esperado', '?')}")
                    st.write(f"✅ Concluídas ({len(resultado['concluidas'])}): {resultado['concluidas'][:20]}{'...' if len(resultado['concluidas']) > 20 else ''}")
                    st.write(f"⏳ Pendentes ({len(resultado['pendentes'])}): {resultado['pendentes'][:20]}{'...' if len(resultado['pendentes']) > 20 else ''}")
                
                st.info("📊 Gerando planilha...")
                excel_buffer, total, conc, pend = gerar_planilha_simplificada(
                    resultado['concluidas'],
                    resultado['pendentes'],
                    scout_name,
                    level,
                    resultado['datas']
                )
                
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                filename = f"{scout_name.replace(' ', '_')}_Simples_{timestamp}.xlsx"
                
                st.session_state.processed = True
                st.session_state.excel_data = excel_buffer
                st.session_state.stats = {
                    'modo': 'simplificado',
                    'total': total,
                    'concluidas': conc,
                    'pendentes': pend
                }
                st.session_state.filename = filename
                
            else:
                st.info("Modo completo ainda não implementado nesta versão. Use o modo simplificado!")

# Resultados
if st.session_state.processed:
    st.markdown("---")
    st.success("🎉 Processamento Concluído!")
    
    stats = st.session_state.stats
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("📊 Total", stats['total'])
    with col2:
        st.metric("✅ Concluídas", stats['concluidas'])
    with col3:
        st.metric("⏳ Pendentes", stats['pendentes'])
    with col4:
        perc = (stats['concluidas']/stats['total']*100) if stats['total']>0 else 0
        st.metric("📈 %", f"{perc:.1f}%")
    
    st.download_button(
        label="⬇️ BAIXAR PLANILHA",
        data=st.session_state.excel_data,
        file_name=st.session_state.filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    if st.button("🔄 Nova Validação"):
        st.session_state.processed = False
        st.rerun()

st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666; padding: 1rem;'>
    <p>🏕️ <b>Scout Progress Validator</b> v4.0</p>
    <p>OCR Simplificado + OCR Completo | Sempre Alerta! ⚜️</p>
</div>
""", unsafe_allow_html=True)
